# -*- coding: utf-8 -*-
"""
Created on Thu Jan  7 13:56:26 2021

@author: USER
"""

import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm, trange

#定義合并函數
def Merge_cells(ws,target_list,start_row,col):
    start = 0
    end = 0
    reference = target_list[0]
    for i in range(len(target_list)):
        if target_list[i] != reference:
            reference=target_list[i]
            end=i-1
            ws.merge_cells(col+str(start + start_row)+":"
                           +col+str(end+start_row))
            start=end+1
        if i==len(target_list)-1:
            end=i
            ws.merge_cells(col+str(start + start_row)+":"
                           +col+str(end+start_row))

#讀取文件和遍歷文件的sheet
wb=load_workbook('D:\成功大學\四下\人工 智慧在中醫上的應用\database\data-O.xlsx')
sheet_names = wb.get_sheet_names()
print(sheet_names[0])

for sheet_name in sheet_names: #遍历每个工作表，抓取数据，并根据要求合并单元格
    ws = wb[sheet_name]
    acupoint_ID_list = [] #針灸ID
    acupoint_name_list = [] #針灸名稱
    disease_ID_list = [] #疾病ID
    disease_name_list = [] #疾病名稱
    meridian_list = [] #經脈
    
    
    for row in tqdm(range(1,ws.max_row-2)):
        acupoint_ID = ws['A' + str(row)].value
        acupoint_name = ws['B' + str(row)].value
        disease_ID = ws['C' + str(row)].value
        disease_name = ws['D'+str(row)].value
        meridian=ws['G'+str(row)].value
       
        acupoint_ID_list.append(acupoint_ID)
        acupoint_name_list.append(acupoint_name)
        disease_ID_list.append(disease_ID)
        disease_name_list.append(disease_name)
        meridian_list.append(meridian)
        
    #调用以上定义的合并单元格函数`Merge_cells`做单元格合并操作    
    start_row=1 #开始行是第一行
    print('start merging...')
    Merge_cells(ws,acupoint_ID_list,start_row,"A")
    print('merging acupoint_ID_list done...')
    Merge_cells(ws,acupoint_name_list,start_row,"B")
    print('merging acupoint_name_list done...')
    Merge_cells(ws,disease_ID_list,start_row,"C")
    print('merging disease_ID_list done...') 
    Merge_cells(ws,disease_name_list, start_row,'D') 
    print('merging disease_name_list done...')
    Merge_cells(ws,meridian_list, start_row,'G')
    print('merging meridian_list done...')

#保存文件    
wb.save('D:\成功大學\四下\人工 智慧在中醫上的應用\database\\finial_data.xlsx')